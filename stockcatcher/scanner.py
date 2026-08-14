import os
import json
import time
import re
import sys
import math
from datetime import datetime, timedelta, timezone, time as dtime
from typing import List, Set, Dict, Any
from dateutil.parser import isoparse
from dotenv import load_dotenv
from openpyxl import load_workbook
from supabase import create_client, Client

# ⚡ 核心：引入 curl_cffi 偽裝 Chrome 瀏覽器，徹底繞過 SSL 憑證驗證失敗與 WAF 阻擋
from curl_cffi import requests as curl_requests
import requests as standard_requests

# 🛡️ 禁用 urllib3 不安全請求警告與載入環境變數
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
load_dotenv()
# 1. 確保連線變數存在
SUPABASE_URL = "https://iatlchzzjkjaetorvvil.supabase.co" #os.getenv("SUPABASE_URL")
# ⚠️ 強烈建議使用 Service Role Key 以繞過 RLS 寫入限制
SUPABASE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImlhdGxjaHp6amtqYWV0b3J2dmlsIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc4NjAzMDQzMywiZXhwIjoyMTAxNjA2NDMzfQ.FckTSOyIo_QCocrgfaGd9mHV2wXRJxSeC5955936hSQ" #os.getenv("SUPABASE_SERVICE_ROLE_KEY")

# 2. 建立全域的 Supabase 客戶端實例
if SUPABASE_URL and SUPABASE_KEY:
    supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
else:
    print("⚠️ [警告] 尚未設定 SUPABASE_URL 或 SUPABASE_SERVICE_ROLE_KEY，資料庫連線失敗。")
    supabase = None

# ============================================================
# ⚙️ [系統配置區] - 實戰參數落鎖 (資安升級：改用環境變數)
# ============================================================
class Config:
    FINMIND_TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJ1c2VyX2lkIjoiRVNCMTc4MTMiLCJlbWFpbCI6Im0yOTk0MDUwOUBob3RtYWlsLmNvbSJ9.iGsA_PLkanve2aATgXU-RD2i7RKOHSLzMEmASMBOcDE"
    FUGLE_API_KEY ="MzJiNjhmNjAtMzRjMy00OGZiLTg3YWQtMTJmMjg3NGE0MDNjIGJlNGVmY2Q2LTE5NDQtNDUzZi1iNTcxLTI5NmIzM2QwOTIzZQ=="
    TELEGRAM_TOKEN = "8480482512:AAGin83kwa61oa5F5rBj4NQMow-C9jsbJug"
    TELEGRAM_CHAT_ID = "-1003613268841"
    TIDE_CONCEPT_JSON = os.getenv("TIDE_CONCEPT_JSON", "config/concepts.json")
    # TIDE 顯示門檻：A / AA / AAA。預設 AA，可避免訊息過長。
    TIDE_MIN_LEVEL = os.getenv("TIDE_MIN_LEVEL", "AA").strip().upper()
    CB_PRIMARY_MARKET_XLSX = os.getenv("CB_PRIMARY_MARKET_XLSX", "")
    CB_PRIMARY_MARKET_KEYWORD = os.getenv("CB_PRIMARY_MARKET_KEYWORD", "CB初級市場資訊")


    # 📊 配額管理
    MAX_POOL_SIZE      = 200
    WARRANT_QUOTA      = 50    
    LISTED_QUOTA       = 100    
    OTC_QUOTA          = 50    
    SCAN_INTERVAL      = 900  

    # 🎯 策略爆發門檻
    ENTRY_MIN_PCT      = 3.5  
    ENTRY_MAX_PCT      = 9.0  
    GRADUATION_PCT     = 9.7  
    VOL_EST_THRESHOLD  = 2.5  

    # ⏰ 時間控制
    MARKET_OPEN        = dtime(9, 0)
    MARKET_CLOSE       = dtime(13, 30)
    AUTO_SHUTDOWN_TIME = dtime(13, 35)  
    RECOVERY_THRESHOLD = dtime(9, 15)  
    CANDLE_TIMEFRAME_MIN  = 5    # 天機圖3K法：K棒週期（分鐘）
    CANDLE_3K_COUNT       = 3    # 天機圖3K法：連續K棒根數
    API_THROTTLE_SLEEP = 1.1  
     
    IS_LOCAL           = False  

# 🗄️ 全域記憶體容器與快照矩陣
stock_info_map = {}  
monitor_data = {}    
finmind_industry_map = {}
global_volume_lookup = {}    # 儲存成交量 (張數)
global_turnover_lookup = {}  # 儲存成交金額 (金流)
last_scan_time = 0
_exchange_map = {}        
_last_fugle_scan = 0.0    
_mis_session = None      
_stocks_with_futures = set()   # 有股票期貨的標的
_stocks_with_cb = set()        # 有可轉債的標的
_cb_primary_market_map = {}    # CB初級市場資訊：sid -> {auction:[], filing:[], board:[]}
_tide_analyzer = None          # TIDE族群共振分析器

# ============================================================
# 🛡️ 🔏 [真．全域落鎖區] 核心工具與通訊函數強制置頂
# ============================================================

def get_now_tw():
    return datetime.now(timezone.utc) + timedelta(hours=8)

def get_previous_trading_day():
    """取得前一個交易日（跳過週末），用於週一盤前無資料時的降級查詢"""
    today = get_now_tw().date()
    prev = today - timedelta(days=1)
    while prev.weekday() >= 5:  # Saturday=5, Sunday=6
        prev -= timedelta(days=1)
    return prev

def is_market_hours():
    now_time = get_now_tw().time()
    return Config.MARKET_OPEN <= now_time <= Config.MARKET_CLOSE

def get_stock_exchange(sid):
    market = stock_info_map.get(sid, {}).get('market', '')
    if market == '上市': return 'tse'
    if market == '上櫃': return 'otc'
    return _exchange_map.get(sid, 'tse')

def wide_ljust(text, width, fillchar=' '):
    text = str(text)
    count = sum(1 for ch in text if ord(ch) > 127)
    return text.ljust(width - count, fillchar)

def get_consumption_badge(rate: float) -> str:
    pct = int(rate * 100)
    if pct >= 80: return f"🟢 {pct}%"
    if pct >= 40: return f"🟡 {pct}%"
    return f"🔴 {pct}%"

def safe_cast(value, target_type, default=0):
    if value is None: return default
    try:
        str_val = str(value).replace(',', '').strip()
        if not str_val: return default
        f = float(str_val)
        if math.isnan(f): return default
        return target_type(f)
    except (ValueError, TypeError):
        return default

def _norm_text(value) -> str:
    return str(value).strip() if value is not None else ""

def _resolve_tide_concepts_path() -> str:
    candidates = [
        Config.TIDE_CONCEPT_JSON,
        os.path.join(os.path.dirname(__file__), Config.TIDE_CONCEPT_JSON),
        os.path.join(os.getcwd(), Config.TIDE_CONCEPT_JSON),
    ]
    for p in candidates:
        if p and os.path.exists(p):
            return p
    return ""

class TideAnalyzer:
    def __init__(self, json_path: str = ""):
        self.concept_groups: Dict[str, List[str]] = {}
        self.stock_to_concepts: Dict[str, List[str]] = {}
        self._load_concepts(json_path or _resolve_tide_concepts_path())

    def _load_concepts(self, json_path: str):
        if not json_path or not os.path.exists(json_path):
            print(f"⚠️ [TIDE] 找不到概念檔 {Config.TIDE_CONCEPT_JSON}，以空字典運行")
            return

        try:
            with open(json_path, "r", encoding="utf-8") as f:
                self.concept_groups = json.load(f)

            for concept_name, stock_list in self.concept_groups.items():
                if not isinstance(stock_list, list):
                    continue
                for stock_id in stock_list:
                    sid = _norm_text(stock_id)
                    if len(sid) != 4 or not sid.isdigit():
                        continue
                    if sid not in self.stock_to_concepts:
                        self.stock_to_concepts[sid] = []
                    self.stock_to_concepts[sid].append(concept_name)

            print(f"✅ [TIDE] 模組加載完成：{len(self.concept_groups)} 個概念、{len(self.stock_to_concepts)} 檔標的")
        except Exception as e:
            print(f"❌ [TIDE] 載入概念 JSON 失敗: {e}")

    def evaluate_resonance(self, target_stock_id: str, snapshot_data: Dict[str, Dict[str, Any]]) -> Dict[str, Any]:
        matched_concepts = self.stock_to_concepts.get(target_stock_id, [])

        if not matched_concepts:
            return {
                "has_concept": False,
                "display_tag": "",
                "resonance_level": "NORMAL",
                "status_text": "",
                "strong_peers_text": "",
                "concept": "",
            }

        best_result = None
        max_resonance_score = -1

        for concept in matched_concepts:
            peer_stocks = self.concept_groups.get(concept, [])
            strong_peers = []

            for sid in peer_stocks:
                if sid == target_stock_id:
                    continue
                info = snapshot_data.get(sid, {})
                change_pct = safe_cast(info.get("change_pct", 0.0), float, 0.0)
                vol_ratio = safe_cast(info.get("vol_ratio", 1.0), float, 1.0)

                # 同族群夥伴：漲幅 >=2% 且 量比 >=1.3 視為同步發動
                if change_pct >= 2.0 and vol_ratio >= 1.3:
                    stock_name = _norm_text(info.get("name", sid))
                    strong_peers.append(f"{sid} {stock_name}")

            resonance_count = len(strong_peers)

            if resonance_count > max_resonance_score:
                max_resonance_score = resonance_count
                if resonance_count >= 3:
                    level = "AAA"
                    status = f"🔥 極強族群共振 (同族群 {resonance_count + 1} 檔同步暴量大漲)"
                elif resonance_count >= 1:
                    level = "AA"
                    status = f"🟢 族群連動發動 (同族群 {resonance_count + 1} 檔同步偏強)"
                else:
                    level = "A"
                    status = "⚠️ 單兵突破 (族群其餘標的尚無明顯跟漲)"

                best_result = {
                    "has_concept": True,
                    "concept": concept,
                    "display_tag": f"{concept} ⚡",
                    "resonance_level": level,
                    "status_text": status,
                    "strong_peers_text": "、".join(strong_peers[:4]) if strong_peers else "無",
                }

        return best_result or {
            "has_concept": False,
            "display_tag": "",
            "resonance_level": "NORMAL",
            "status_text": "",
            "strong_peers_text": "",
            "concept": "",
        }

def _build_tide_snapshot(current_sid: str, current_up_pct: float, current_ratio: float) -> Dict[str, Dict[str, Any]]:
    snapshot = {}
    for sid, info in stock_info_map.items():
        data = monitor_data.get(sid, {})
        snapshot[sid] = {
            "name": info.get("name", sid),
            "change_pct": safe_cast(data.get("last_up_pct", 0.0), float, 0.0),
            "vol_ratio": safe_cast(data.get("last_ratio", 1.0), float, 1.0),
        }

    if current_sid:
        if current_sid not in snapshot:
            snapshot[current_sid] = {"name": current_sid, "change_pct": 0.0, "vol_ratio": 1.0}
        snapshot[current_sid]["change_pct"] = safe_cast(current_up_pct, float, 0.0)
        snapshot[current_sid]["vol_ratio"] = safe_cast(current_ratio, float, 1.0)

    return snapshot

def _tide_level_rank(level: str) -> int:
    mapping = {"A": 1, "AA": 2, "AAA": 3}
    return mapping.get(_norm_text(level).upper(), 0)

def build_tide_monitor_message(sid: str, up_pct: float, ratio: float) -> str:
    if _tide_analyzer is None:
        return ""

    try:
        snapshot = _build_tide_snapshot(sid, up_pct, ratio)
        tide = _tide_analyzer.evaluate_resonance(sid, snapshot)
        if not tide.get("has_concept", False):
            return ""

        current_level = tide.get("resonance_level", "A")
        # 觸發門檻開關：僅在達到 TIDE_MIN_LEVEL 時才顯示 TIDE 區塊
        if _tide_level_rank(current_level) < _tide_level_rank(Config.TIDE_MIN_LEVEL):
            return ""

        concept = tide.get("concept", "")
        display_tag = tide.get("display_tag", concept)
        status_text = tide.get("status_text", "")
        peers = tide.get("strong_peers_text", "無")
        
        # 對 peers 進行格式化：如果包含多個標的，按逗號分割並添加折行縮排
        if "," in peers and len(peers) > 40:
            peers_list = [p.strip() for p in peers.split(",") if p.strip()]
            peers_formatted = ",\n                    ".join(peers_list)
        else:
            peers_formatted = peers

        return (
            f"🔖 *細分族群：* {display_tag}\n"
            f"    🌊 *【TIDE 族群共振監控】*\n"
            f"         🔹 概念題材：\n               {concept}\n"
            f"         🔹 共振狀態：\n               {status_text}\n"
            f"         🔹 同步發動：\n               {peers_formatted}"
        )
    except Exception:
        return ""

def _format_ymd(value) -> str:
    if value is None:
        return "-"
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    txt = _norm_text(value)
    # 常見格式：2026-07-27 00:00:00 / 2026/07/27
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(txt, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    parsed = _extract_yyyy_mm_dd_from_text(txt)
    if parsed:
        return parsed.strftime('%Y-%m-%d')
    return txt or "-"

def _pick_value_by_keys(row: dict, keys: List[str]) -> str:
    for k, v in row.items():
        key = _norm_text(k).lower()
        if any(token in key for token in keys):
            val = _norm_text(v)
            if val:
                return val
    return ""

def _resolve_cb_primary_xlsx_path() -> str:
    # 1) 若有明確指定完整檔名/路徑，優先使用
    explicit_candidates = [
        Config.CB_PRIMARY_MARKET_XLSX,
        os.path.join(os.path.dirname(__file__), Config.CB_PRIMARY_MARKET_XLSX) if Config.CB_PRIMARY_MARKET_XLSX else "",
        os.path.join(os.getcwd(), Config.CB_PRIMARY_MARKET_XLSX) if Config.CB_PRIMARY_MARKET_XLSX else "",
    ]
    for p in explicit_candidates:
        if p and os.path.exists(p):
            return p

    # 2) 未指定或指定不存在時，改用關鍵字搜尋 xlsx（不依賴日期）
    keyword = _norm_text(Config.CB_PRIMARY_MARKET_KEYWORD)
    search_dirs = [os.path.dirname(__file__), os.getcwd()]
    matches = []

    for base_dir in search_dirs:
        try:
            for name in os.listdir(base_dir):
                if not name.lower().endswith('.xlsx'):
                    continue
                if keyword and keyword not in name:
                    continue
                full_path = os.path.join(base_dir, name)
                if os.path.isfile(full_path):
                    matches.append(full_path)
        except Exception:
            pass

    if matches:
        # 同關鍵字多檔時選擇最新更新檔
        matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
        return matches[0]

    return ""

def _classify_cb_stage(row: dict) -> str:
    # 透過欄名與欄值綜合判斷階段
    blob = " ".join([_norm_text(k) for k in row.keys()] + [_norm_text(v) for v in row.values()]).lower()
    if ('詢圈' in blob) or ('竞拍' in blob) or ('競拍' in blob):
        return 'auction'
    if '送件' in blob:
        return 'filing'
    if ('董事會' in blob and '通過' in blob) or ('發行通過' in blob):
        return 'board'
    return ''

def load_cb_primary_market_map_from_xlsx() -> dict:
    path = _resolve_cb_primary_xlsx_path()
    if not path:
        print(f"⚠️ [CB初級市場] 找不到檔案 {Config.CB_PRIMARY_MARKET_XLSX}，略過整合。")
        return {}

    result = {}
    try:
        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        if ws is None:
            return {}
        active_stage = ''
        headers = []

        for row_vals in ws.iter_rows(values_only=True):
            vals = [_norm_text(v) for v in row_vals]
            joined = " ".join(v for v in vals if v)
            if not joined:
                continue

            # 區塊起始（你的Excel是多區段表）
            if '詢圈/競拍 標的' in joined:
                active_stage = 'auction'
                headers = []
                continue
            if '送件標的' in joined:
                active_stage = 'filing'
                headers = []
                continue
            if ('董事會通過發行標的' in joined) or ('董事會發行通過' in joined):
                active_stage = 'board'
                headers = []
                continue

            # 各區塊的欄名列
            if 'CB代碼' in joined and '標的名稱' in joined:
                headers = vals
                continue

            if not active_stage or not headers:
                continue

            row = {}
            has_payload = False
            for i, h in enumerate(headers):
                h = _norm_text(h)
                if not h:
                    continue
                val = row_vals[i] if i < len(row_vals) else None
                row[h] = val
                if _norm_text(val):
                    has_payload = True

            if not has_payload:
                continue

            sid = _extract_sid_by_priority(
                row,
                ['發行公司代號', '公司代號', '證券代號', '股票代號', '標的代號', 'SecuritiesCompanyCode', 'UnderlyingStockID']
            )
            if not sid:
                cb_code = _pick_value_by_keys(row, ['cb代碼', 'cb code', '債券代號'])
                cb_code = _norm_text(cb_code)
                if len(cb_code) >= 4 and cb_code[:4].isdigit() and cb_code[:4] != '0000':
                    sid = cb_code[:4]

            if not sid:
                continue

            name = _pick_value_by_keys(row, ['標的名稱', '名稱', '公司簡稱', '證券名稱', '發行公司']) or sid
            list_date = _pick_value_by_keys(row, ['掛牌日', '掛牌日期', '上市日'])
            conv_price = _pick_value_by_keys(row, ['發行轉換價格', '轉換價格', '轉換價'])
            filing_date = _pick_value_by_keys(row, ['送件日', '送件日期'])
            announce_date = _pick_value_by_keys(row, ['公告日', '公告日期'])

            if sid not in result:
                result[sid] = {'auction': [], 'filing': [], 'board': []}

            if active_stage == 'auction':
                result[sid]['auction'].append({
                    'name': name,
                    'list_date': _format_ymd(list_date),
                    'conv_price': conv_price or '-'
                })
            elif active_stage == 'filing':
                result[sid]['filing'].append({
                    'name': name,
                    'filing_date': _format_ymd(filing_date)
                })
            elif active_stage == 'board':
                result[sid]['board'].append({
                    'name': name,
                    'announce_date': _format_ymd(announce_date)
                })

        print(f"✅ [CB初級市場] 載入完成：{len(result)} 檔股票具備CB階段資訊。")
    except Exception as e:
        print(f"❌ [CB初級市場] 讀檔失敗: {e}")
        return {}

    return result

def build_cb_primary_market_message(sid: str) -> str:
    info = _cb_primary_market_map.get(sid)
    if not info:
        return ""

    lines = ["🧾 *CB初級市場：*"]

    if info.get('auction'):
        item = info['auction'][0]
        lines.append(f"     🔸 詢圈/競拍：{item['name']}\n               掛牌 {item['list_date']} | 轉換價 {item['conv_price']}")

    if info.get('filing'):
        item = info['filing'][0]
        lines.append(f"     🔸 送件標的：{item['name']}\n               送件日 {item['filing_date']}")

    if info.get('board'):
        item = info['board'][0]
        lines.append(f"     🔸 董事會通過：{item['name']}\n               公告日 {item['announce_date']}")

    return "\n".join(lines) if len(lines) > 1 else ""
# ---------------------------------------------------------
# [通訊層模組] LINE Bot API 推播功能
# ---------------------------------------------------------

def _build_flex_row(label: str, value: str, color: str = "#334155", weight: str = "regular") -> dict:
    """[輔助函數] 構建 LINE Flex Message 單列，優化長標籤排版比例"""
    return {
        "type": "box",
        "layout": "baseline",
        "spacing": "sm",
        "contents": [
            {"type": "text", "text": label, "color": "#64748b", "size": "sm", "flex": 3},
            {"type": "text", "text": str(value), "color": color, "weight": weight, "size": "sm", "wrap": True, "flex": 7}
        ]
    }

def send_line_flex_warrant_alert(
    sid: str, name: str, strategy: str, lp: float, pct_str: str, up_pct: float, 
    ratio: float, stop_loss: float, pressure_digestion: str, energy_slope: str, 
    fut_flag: str, cb_flag: str, industry: str, time_str: str
):
    """
    [通訊層] 專屬權證主力的 LINE Flex Message，帶有 nstock 走勢超連結
    """
    # 🛡️ 安全規範：透過環境變數讀取憑證，絕不硬編碼
    token = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
    target_id = os.getenv("LINE_GROUP_ID")

    if not token or not target_id:
        print("❌ [錯誤] LINE 憑證遺失，請檢查 .env 檔案配置。")
        return

    # 動態判定漲跌顏色
    price_color = "#dc2626" if up_pct > 0 else "#16a34a" if up_pct < 0 else "#475569"
    nstock_url = f"https://www.nstock.tw/stock_info?stock_id={sid}"

    flex_payload = {
        "type": "flex",
        "altText": f"🔥 權證主力進場: {sid} {name}",
        "contents": {
            "type": "bubble",
            "size": "mega",
            "header": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#7c2d12",
                "contents": [
                    {"type": "text", "text": "🔥 權證主力 - 籌碼共振訊號", "color": "#ffffff", "weight": "bold", "size": "sm"}
                ]
            },
            "body": {
                "type": "box",
                "layout": "vertical",
                "spacing": "md",
                "contents": [
                    {
                        "type": "text",
                        "text": f"{sid} {name}",
                        "weight": "bold",
                        "size": "xl",
                        "color": "#1d4ed8",
                        "decoration": "underline",
                        "action": {
                            "type": "uri",
                            "label": "查看 nstock",
                            "uri": nstock_url
                        }
                    },
                    {
                        "type": "box",
                        "layout": "vertical",
                        "spacing": "sm",
                        "contents": [
                            _build_flex_row("策略", strategy),
                            _build_flex_row("現價", f"{lp} ({pct_str})", color=price_color, weight="bold"),
                            _build_flex_row("量能比", f"{ratio}x"),
                            _build_flex_row("停損價", stop_loss),
                            _build_flex_row("壓力消化", str(pressure_digestion)),
                            _build_flex_row("能量斜率", str(energy_slope)),
                            _build_flex_row("衍生品", f"股期 {fut_flag} | 可轉債 {cb_flag}"),
                            _build_flex_row("產業別", str(industry))
                        ]
                    }
                ]
            },
            "footer": {
                "type": "box",
                "layout": "vertical",
                "contents": [
                    {"type": "text", "text": f"觸發時間: {time_str}", "size": "xs", "color": "#94a3b8", "align": "end"}
                ]
            }
        }
    }

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    try:
        res = standard_requests.post(url, headers=headers, json={"to": target_id, "messages": [flex_payload]}, timeout=5)
        if res.status_code != 200:
            print(f"⚠️ [LINE] 推播失敗 (HTTP {res.status_code}): {res.text}")
    except Exception as e:
        print(f"❌ [LINE] 發送異常: {e}")

def send_line_status_flex_message(version: str, mode: str, time_str: str):
    """
    [通訊層] 發送系統狀態專用的 LINE Flex Message 卡片
    """
    token = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
    target_id = os.getenv("LINE_GROUP_ID")

    if not token or not target_id:
        print("⚠️ [LINE] Token 或 Group ID 遺失，略過系統狀態通知。")
        return

    # 構建科技感的系統狀態卡片
    flex_payload = {
        "type": "flex",
        "altText": f"🚀 系統啟動成功 ({version})",
        "contents": {
            "type": "bubble",
            "size": "kilo", # 狀態通知不需要太大，使用 kilo 尺寸即可
            "header": {
                "type": "box",
                "layout": "vertical",
                "backgroundColor": "#0f172a", # 沉穩的深藍色，與股票警報區隔
                "contents": [
                    {"type": "text", "text": "🖥️ 系統狀態報告", "color": "#ffffff", "weight": "bold", "size": "sm"}
                ]
            },
            "body": {
                "type": "box",
                "layout": "vertical",
                "spacing": "md",
                "contents": [
                    {
                        "type": "text",
                        "text": "比鼻的天機選股",
                        "weight": "bold",
                        "size": "xl",
                        "color": "#1e293b"
                    },
                    {
                        "type": "text",
                        "text": "🟢 雷達上線運作中",
                        "size": "md",
                        "color": "#16a34a",
                        "weight": "bold"
                    },
                    {
                        "type": "separator",
                        "margin": "md"
                    },
                    {
                        "type": "box",
                        "layout": "vertical",
                        "spacing": "sm",
                        "contents": [
                            # 沿用我們先前寫好的 _build_flex_row 輔助函式
                            _build_flex_row("版本", version, weight="bold"),
                            _build_flex_row("模式", mode),
                            _build_flex_row("時間", time_str)
                        ]
                    }
                ]
            }
        }
    }

    url = "https://api.line.me/v2/bot/message/push"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {token}"
    }
    
    try:
        res = standard_requests.post(url, headers=headers, json={"to": target_id, "messages": [flex_payload]}, timeout=5)
        if res.status_code != 200:
            print(f"⚠️ [LINE] 系統通知發送失敗 (HTTP {res.status_code}): {res.text}")
    except Exception as e:
        print(f"❌ [LINE] 系統通知連線異常: {e}")
# ==========================================
# 1. 策略名稱標準化邏輯 (Data Normalization)
# ==========================================
def clean_category_name(raw_name: str) -> str:
   """
   將爬蟲抓取的中文策略名稱，轉換為標準英文代碼 (Slug)。
   此舉確保資料庫欄位的一致性，利於後續建立 Index 與 API 查詢。
   """
   if not raw_name:
       return "unknown_strategy"
       
   # 防禦性處理：去除空白並轉小寫，避免「3K 突破」或「3k突破」的差異導致判斷失敗
   clean_name = str(raw_name).replace(" ", "").lower()
   
   # 放寬條件，擷取核心關鍵字進行映射
   if "權證" in clean_name:
       return "warrant_3k"
   elif "3k" in clean_name:
       return "volume_3k"
       
   return "unknown_strategy"
    
def send_tg_alert(sid, strategy_name, lp, high=0.0, low=0.0, ratio=0.0, up_pct=0.0):
    if up_pct > 9.0:
        return
       
    info = stock_info_map.get(sid, {})
    data = monitor_data.get(sid, {})

    # 每策略獨立 CD 鎖，避免短時間內重複洗頻
    alert_times = data.setdefault('last_alert_by_strategy', {})
    if time.time() - alert_times.get(strategy_name, 0) < 1800:
        return
       
    badge = "🎫 [權證主力標的] " if info.get('is_protected') else f"[{info.get('market', '未知')}] "
    scenario = "🚨 [訊號觸發]"
    consumption_str = get_consumption_badge(data.get('last_consumption', 0.45))
    is_acc = data.get('is_accelerating', False)
   
    # 確立防守點位
    stop_loss_price = low if low > 0 else lp
   
    pct_arrow = "📈" if up_pct >= 0 else "📉"
    pct_str = f"+{up_pct}%" if up_pct > 0 else f"{up_pct}%"
   
    # 衍生商品 Flag 狀態轉換 (✅ / ❌)
    futures_flag = "✅" if sid in _stocks_with_futures else "❌"
    cb_flag = "✅" if sid in _stocks_with_cb else "❌"
    cb_primary_block = build_cb_primary_market_message(sid) if cb_flag == "✅" else ""
    cb_primary_line = f"{cb_primary_block}\n" if cb_primary_block else ""
    tide_block = build_tide_monitor_message(sid, up_pct, ratio)
    tide_line = f"{tide_block}\n" if tide_block else ""
   
    # 1. 建立基礎訊息區塊 (去除所有末尾的 \n，由程式統一處理)
    msg_lines = [
    f"{scenario}",
    f"🎯 *核心策略：* {badge}{strategy_name}",
    "━━━━━━━━━━━━",
    f"📈 *標的：* [{sid} {info.get('name', '')}](https://www.nstock.tw/stock_info?stock_id={sid})",
    f"💰 *現價：* `{lp}` {pct_arrow} `{pct_str}`",
    f"📊 *預估量比：* `{ratio}x`",
    f"📐 *3K高位：* `{data.get('high', 0.0)}`",
    f"🛡️ *策略停損：* `{stop_loss_price}`",
    f"💥 *壓力消化：* {consumption_str}",
    f"🚀 *能量斜率：* {'陡增' if is_acc else '平穩'}",
    f"📦 *衍生品：* 股期 {futures_flag} | CB {cb_flag}",
    f"🏷️ *產業類別：* `{info.get('industry', '未知產業')}`"
   ]

    # 2. 動態擴展區塊：CB 與 TIDE 資訊
    # 使用 strip() 確保清除前後多餘的空白或換行。若為空，則不會加入 msg_lines。

    if cb_primary_line and str(cb_primary_line).strip():
      msg_lines.append(str(cb_primary_line).strip())

    if tide_line and str(tide_line).strip():
       msg_lines.append(str(tide_line).strip())

    # 3. 結尾區塊
    msg_lines.extend([
    "━━━━━━━━━━━━",
    f"⏰ {get_now_tw().strftime('%H:%M:%S')}"
    ])

    # 4. 完美組合：將陣列轉換為單一字串，並在每行之間插入換行符號
    msg = "\n".join(msg_lines)
   
    url = f"https://api.telegram.org/bot{Config.TELEGRAM_TOKEN}/sendMessage"
   
    try:
        standard_requests.post(
            url,
            json={
                "chat_id": Config.TELEGRAM_CHAT_ID,
                "text": msg,
                "parse_mode": "Markdown"
            },
            timeout=5
        )
        alert_times[strategy_name] = time.time()
    except Exception as e:
        print(f"[錯誤] Telegram 發送通知失敗: {e}")
    # ==========================================
    # 🛡️ 【架構師防護網】阻擋測試訊號污染資料庫
    # ==========================================
    # 請根據你每日測試推播的實際特徵進行調整。
    # 這裡假設你的測試訊號會在 strategy_name 包含 "測試" 兩字，或是帶有特定的測試股號
    if "測試" in strategy_name or "test" in strategy_name.lower() or sid == "0000":
        print(f"⚠️ [DB 寫入略過] 偵測到系統測試訊號 ({strategy_name})，不寫入正式資料庫。")
        return  # 提早返回，攔截下方的 DB 寫入動作
    # 檔案位置：scanner.py (send_tg_alert 函式內，防衛語句之下)

    try:
        # ==========================================
        # 🎯 【資料清洗與提取】安全的擴充欄位取值
        # ==========================================
        cb_info = info.get("cb_info", {})
        tide_info = info.get("tide_info", {})

        # ==========================================
        # 📦 【Payload 組裝】準備寫入 Supabase JSONB
        # ==========================================
        stock_payload = {
            "stock_id": str(sid),
            "stock_name": info.get("name", ""),
            "price": float(lp),
            "pct": float(up_pct),
            "vol_ratio": float(ratio),
            "stop_loss": float(stop_loss_price),
            "industry": info.get("industry", "-"),
            "sub_industry": info.get("sub_industry", "-"),
            "3k_high": float(data.get("high", 0.0)),
            "pressure_digestion": data.get("last_consumption", "0%"),
            "energy_slope": "陡增" if is_acc else "平穩",
            "derivatives": f"股期 {futures_flag} | CB {cb_flag}",
            "cb_info": cb_info,
            "tide_info": tide_info,
        }
        full_strategy_string = f"{badge}{strategy_name}"
        standard_category = clean_category_name(full_strategy_string)
        db_payload = {
            "category": standard_category,
            "data": stock_payload,
        }
        supabase.table("tianji_signals").insert(db_payload).execute()
        print(f"✅ [DB 寫入成功] 策略: {strategy_name} | 標的: {sid} {info.get('name', '')}")
    except Exception as e:
        print(f"❌ [DB 寫入前置轉換或連線失敗] 錯誤: {e}")

# ==========================================
# 檔案位置：scanner.py (工具函式定義區塊)
# ==========================================
def save_signal_to_supabase(category: str, stock_data: dict) -> None:
    """
    將單一標的訊號寫入 Supabase (Append-only 模式)
    包含完整的防呆與例外處理，絕不干擾主執行緒。
    """
    if not supabase:
        print("❌ [DB 跳過寫入] Supabase 客戶端未初始化。")
        return

    try:
        # 確保寫入的欄位符合我們定義的 Schema
        row_payload = {
            "category": category,
            "data": stock_data  # 直接將字典寫入 JSONB 欄位
        }

        # 執行寫入 (created_at 會由 Supabase 伺服器端自動產生)
        response = supabase.table("tianji_signals").insert(row_payload).execute()
        print(f"✅ [DB 寫入成功] {category} | {stock_data.get('stock_id')} {stock_data.get('stock_name')}")

    except Exception as e:
        # 發生錯誤時僅印出 Log，不使用 raise 拋出異常，確保 TG 通知繼續執行
        print(f"❌ [DB 寫入失敗] {category} | 錯誤原因: {e}")

def perform_strategy_test():
    print("📡 啟動自動化測試：發送驗證訊號...", flush=True)
    sid = "2313"
    stock_info_map[sid] = {'name': '華通(測試標的)', 'market': '上市', 'is_protected': False, 'industry': '電子零組件業'}
    _stocks_with_futures.add(sid)
    monitor_data[sid] = {
        'last_alert_time': 0, 'last_up_pct': 0.0, 'last_consumption': 0.85,
        'is_accelerating': False, 'history_prices': [250.0], 'high': 245.0, 'low': 233.0
    }
    send_tg_alert(sid, "🔥 策略二：3K突破+量能異常測試", 250.0, 245.0, 233.0, 1.8, 0.0)
    del stock_info_map[sid]
    del monitor_data[sid]
    _stocks_with_futures.remove(sid)
    
    # 1. 取得當前台灣時間 (UTC+8)
    tw_now = datetime.now(timezone.utc) + timedelta(hours=8)
    time_str = tw_now.strftime('%Y-%m-%d %H:%M:%S')
    # 2. 準備啟動廣播訊息
    startup_msg = f"🚀 系統啟動成功\n天機選股雷達上線運作中！\n時間：{time_str}\n模式：TG & LINE 雙軌監控"
    print(startup_msg)
    send_line_status_flex_message("v2.3", "Telegram & LINE 雙軌監控", time_str)
    print("✅ 自動化測試驗證訊號已成功送出！")

def should_exclude(sid, name, industry):
    sid, name, industry = str(sid).strip(), str(name), str(industry)
    if sid.startswith('00') or sid.startswith('01') or sid.startswith('03') or \
       any(k in name for k in ["ETF", "受益憑證", "基金", "指數", "債券", "存託憑證"]) or \
       any(k in industry for k in ["ETF", "受益憑證", "指數", "債券", "存託憑證"]): return True
    if sid.startswith('28') or sid.startswith('58') or sid.startswith('60') or \
       any(k in name for k in ["金控", "銀行", "保險", "證券", "人壽", "信託", "期貨"]) or \
       any(k in industry for k in ["金融", "保險", "證券", "金控"]): return True
    return False

def fetch_disposition_stocks():
    disposition_set = set()
    try:
        url_notice = "https://openapi.twse.com.tw/v1/announcement/notice"
        res_notice = curl_requests.get(url_notice, impersonate="chrome120", timeout=10).json()
        for item in res_notice:
            sid = item.get('證券代號', '').strip()
            if sid: disposition_set.add(sid)
    except Exception: pass
    return disposition_set

def fetch_mis_batch_all():
    global _mis_session
    if _mis_session is None:
        _mis_session = curl_requests.Session()
        try: _mis_session.get("https://mis.twse.com.tw/stock/index.jsp", impersonate="chrome120", timeout=6)
        except: pass
       
    results = {}
    all_sids = list(stock_info_map.keys())
    if not all_sids: return results
   
    BATCH_SIZE = 25  
    for i in range(0, len(all_sids), BATCH_SIZE):
        batch = all_sids[i:i + BATCH_SIZE]
        ex_ch = '|'.join(f"{get_stock_exchange(s)}_{s}.tw" for s in batch)
        url = f"https://mis.twse.com.tw/stock/api/getStockInfo.jsp?ex_ch={ex_ch}&_={int(time.time()*1000)}"
        try:
            res = _mis_session.get(url, impersonate="chrome120", timeout=6).json()
            for item in res.get('msgArray', []):
                sid = item.get('c', '').strip()
                if not sid: continue
                z, y = item.get('z', '-'), safe_cast(item.get('y', '0'), float)
                lp = y if (z in ('-', '0') or safe_cast(z, float) <= 0) else safe_cast(z, float)
                if lp <= 0: continue
                v = safe_cast(item.get('v', '0').replace(',', ''), int)  
                f_str = item.get('f', '')
                ask3 = sum(safe_cast(x, int) for x in f_str.split('_')[:3]) if f_str else 0
                up_pct = round((lp - y) / y * 100, 2) if y > 0 else 0.0
                is_traded = z not in ('-', '0') and safe_cast(z, float) > 0
                results[sid] = {'lp': lp, 'v': v, 'ask3': ask3, 'up_pct': up_pct, 'is_traded': is_traded}
        except Exception: pass
    return results

_TWSE_INDUSTRY_CODE_MAP = {
    '01': '水泥工業',    '02': '食品工業',    '03': '塑膠工業',    '04': '紡織纖維',
    '05': '電機機械',    '06': '電器電纜',    '07': '化學生技醫療', '08': '玻璃陶瓷',
    '09': '造紙工業',    '10': '鋼鐵工業',    '11': '橡膠工業',    '12': '汽車工業',
    '13': '電子工業',    '14': '建材營造',    '15': '航運業',      '16': '觀光餐旅',
    '17': '金融保險',    '18': '貿易百貨',    '19': '綜合',        '20': '其他',
    '21': '化學工業',    '22': '生技醫療業',  '23': '油電燃氣業',  '24': '半導體業',
    '25': '電腦及週邊設備業', '26': '光電業', '27': '通信網路業',  '28': '電子零組件業',
    '29': '電子通路業',  '30': '資訊服務業',  '31': '其他電子業',  '32': '文化創意業',
    '33': '農業科技業',  '34': '電子商務',    '35': '綠能環保',    '36': '數位雲端',
    '37': '運動休閒',    '38': '居家生活',    '39': '管理股票',
}

def fetch_finmind_industry_mapping():
    mapping = {
        "2330": {"name": "台積電", "industry": "半導體業"}, "2317": {"name": "鴻海", "industry": "其他電子業"},
        "2454": {"name": "聯發科", "industry": "半導體業"}, "2382": {"name": "廣達", "industry": "電腦週邊業"}
    }
    try:
        twse_profile = curl_requests.get("https://openapi.twse.com.tw/v1/opendata/t187ap03_L", impersonate="chrome120", timeout=10).json()
        if isinstance(twse_profile, list):
            for row in twse_profile:
                sid = row.get('公司代號', '').strip()
                if len(sid) == 4:
                    ind_code = row.get('產業別', '').strip()
                    ind_name = _TWSE_INDUSTRY_CODE_MAP.get(ind_code, '上市其他')
                    mapping[sid] = {"name": row.get('公司簡稱', '').strip(), "industry": ind_name}
    except Exception: pass
    try:
        from FinMind.data import DataLoader
        dl = DataLoader()
        try:
            if Config.FINMIND_TOKEN: dl.login_by_token(Config.FINMIND_TOKEN)
        except Exception: pass
        df_info = dl.taiwan_stock_info()
        df_info['_sid'] = df_info['stock_id'].astype(str).str.strip()
        tpex_df = df_info[(df_info['type'] == 'tpex') & (df_info['_sid'].str.len() == 4)]
        for _, row in tpex_df.iterrows():
            sid = row['_sid']
            mapping[sid] = {"name": str(row.get('stock_name', '')).strip(), "industry": str(row.get('industry_category', '上櫃其他')).strip()}
    except Exception: pass
    return mapping

def fetch_market_candidates(market_type="上市"):
    candidates = []
    disposition_set = fetch_disposition_stocks()
    try:
        if market_type == "上市":
            url = "https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL"
            res = curl_requests.get(url, impersonate="chrome120", timeout=10).json()
            rows = res if isinstance(res, list) else []
            for i in rows:
                sid = i.get('Code')
                if not sid or len(sid) != 4 or sid in disposition_set or sid in stock_info_map: continue
                fm = finmind_industry_map.get(sid, {"name": i.get('Name', f"上市_{sid}").strip(), "industry": "電子零組件業"})
                if should_exclude(sid, fm['name'], fm['industry']): continue
                vol = safe_cast(i.get('TradeVolume'), int) // 1000  
                turnover = safe_cast(i.get('TradeValue'), float)
                candidates.append({'sid': sid, 'up_pct': 0.0, 'vol': vol, 'turnover': turnover, 'market': '上市', 'name': fm['name'], 'ind': fm['industry']})
        else:
            url = "https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes"
            res = curl_requests.get(url, impersonate="chrome120", timeout=10).json()
            rows = res if isinstance(res, list) else []
            for i in rows:
                sid = i.get('SecuritiesCompanyCode', '').strip()
                if not sid or len(sid) != 4 or sid in disposition_set or sid in stock_info_map: continue
                raw_name = i.get('CompanyName', i.get('SecuritiesCompanyName', f"上櫃_{sid}")).strip()
                fm = finmind_industry_map.get(sid, {"name": raw_name, "industry": "半導體業"})
                if should_exclude(sid, fm['name'], fm['industry']): continue
                vol = safe_cast(i.get('TradingShares'), int) // 1000  
                turnover = safe_cast(i.get('TradingAmount'), float)
                candidates.append({'sid': sid, 'up_pct': 0.0, 'vol': vol, 'turnover': turnover, 'market': '上櫃', 'name': fm['name'], 'ind': fm['industry']})
               
        if candidates:
            candidates.sort(key=lambda x: x['turnover'], reverse=True)
            for idx, c in enumerate(candidates): c['rank_t'] = idx
            candidates.sort(key=lambda x: x['vol'], reverse=True)
            for idx, c in enumerate(candidates): c['rank_v'] = idx
            candidates.sort(key=lambda x: x['rank_t'] + x['rank_v'])
    except Exception: pass
    return candidates

def refresh_pool_v90():
    global last_scan_time
    now = time.time()
    _non_protected = [s for s in stock_info_map if not stock_info_map[s].get('is_protected')]
    force_fill = (len(_non_protected) == 0 and len(global_volume_lookup) > 0)
   
    if get_now_tw().time() < dtime(9, 45) and not force_fill:
        return
    if now - last_scan_time < Config.SCAN_INTERVAL and len(stock_info_map) >= 120: return
    last_scan_time = now
   
    market_configs = [
        {"type": "上市", "quota": Config.LISTED_QUOTA, "fetch_func": lambda: fetch_market_candidates("上市")},
        {"type": "上櫃", "quota": Config.OTC_QUOTA, "fetch_func": lambda: fetch_market_candidates("上櫃")}
    ]

    for config in market_configs:
        m_type, m_quota = config["type"], config["quota"]
        candidates = config["fetch_func"]()
       
        current_dynamic_sids = [s for s in stock_info_map if stock_info_map[s]['market'] == m_type and not stock_info_map[s].get('is_protected')]
        to_remove_count = max(0, len(current_dynamic_sids) + len(candidates) - m_quota)
        to_remove_list = list(set(current_dynamic_sids))[:to_remove_count]
       
        for rsid in to_remove_list:
            stock_info_map.pop(rsid, None)
            monitor_data.pop(rsid, None)

        remaining = len([s for s in stock_info_map if stock_info_map[s]['market'] == m_type and not stock_info_map[s].get('is_protected')])
        vacancy = m_quota - remaining

        for cand in candidates[:vacancy]:
            csid = cand['sid']
            stock_info_map[csid] = {'name': cand['name'], 'market': m_type, 'is_protected': False, 'industry': cand['ind']}
            monitor_data[csid] = {
                "high": 0.0, "low": 9999.0, "y_vol": cand['vol'],
                "trig_both": False, "trig_3k": False, "trig_vol": False, "trig_策略四": False,
                "state": 0, "point_a": -1.0, "point_b": 9999.0,  
                "last_alert_time": 0, "last_up_pct": 0.0, "last_ratio": 1.0, "last_consumption": 0.0, "is_accelerating": False, "history_prices": [],
                "candle_window": [], "last_alert_by_strategy": {}
            }

# ============================================================
# 🕵️♂️ ✅ [V2.3 真．衍生商品與無結構權證解析] 導入 curl_cffi 與限流防禦
# ============================================================

def _extract_sids_from_taifex_rows(rows: List, keys_to_check: List[str]) -> Set[str]:
    extracted = set()
    if not isinstance(rows, list): return extracted
    for row in rows:
        for key in keys_to_check:
            val = str(row.get(key, '')).strip()
            if len(val) == 4 and val.isdigit():
                extracted.add(val)
                break
    return extracted

def _extract_yyyy_mm_dd_from_text(raw: str):
    """將常見日期字串轉為 date，支援民國/西元與分隔符。"""
    txt = str(raw).strip()
    if not txt:
        return None

    digits = ''.join(ch for ch in txt if ch.isdigit())
    if not digits:
        return None

    try:
        # 民國 1130105
        if len(digits) == 7:
            y = int(digits[:3]) + 1911
            m = int(digits[3:5])
            d = int(digits[5:7])
            return datetime(y, m, d).date()
        # 西元 20260105
        if len(digits) == 8:
            y = int(digits[:4])
            m = int(digits[4:6])
            d = int(digits[6:8])
            return datetime(y, m, d).date()
    except Exception:
        return None

    return None

def _extract_sid_by_priority(row: dict, preferred_keys: List[str]) -> str:
    """優先從指定欄位抓 4 碼標的代號，避免全欄位暴力掃描造成誤判。"""
    for key in preferred_keys:
        val = str(row.get(key, '')).strip()
        if len(val) == 4 and val.isdigit() and val != '0000':
            return val
    return ""

FUTURES_CACHE_FILE = "stock_futures_cache.json"

def _load_futures_from_cache(max_age_days: int = 1) -> Set[str]:
    if not os.path.exists(FUTURES_CACHE_FILE):
        return set()

    try:
        with open(FUTURES_CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)

        cache_time = datetime.fromisoformat(data.get("timestamp", "2000-01-01"))
        today = get_now_tw()
        if (today - cache_time).days <= max_age_days:
            return set(data.get("sids", []))
    except Exception as e:
        print(f"⚠️ [股期快取讀取失敗] {e}")

    return set()

def _save_futures_to_cache(futures_set: Set[str]):
    try:
        data = {
            "timestamp": get_now_tw().isoformat(),
            "sids": list(futures_set)
        }
        with open(FUTURES_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ [股期快取寫入失敗] {e}")

def fetch_stock_futures_set():
    print("📡 正在調閱全市場股票期貨成份股清單 (三層防禦模式)...", flush=True)
    combined_set = set()
    
    # 提前建立本系統可辨識的現股母體 (過濾雜訊用)
    equity_universe = set(global_volume_lookup.keys()) | set(finmind_industry_map.keys())
    
    stealth_headers = {
        "Accept": "application/json, text/html, */*",
        "Accept-Language": "zh-TW,zh;q=0.9,en-US;q=0.8,en;q=0.7",
        "Connection": "keep-alive",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }

    # 🛑 [策略一] OpenAPI JSON 解析
    try:
        res = curl_requests.get(
            "https://openapi.taifex.com.tw/v1/StockFutures", 
            impersonate="chrome120", 
            headers=stealth_headers, 
            verify=False,
            timeout=10
        )
        # 確保回傳的是 JSON 而不是 HTML 阻擋頁面
        if res.status_code == 200 and res.text.strip().startswith("["):
            found = _extract_sids_from_taifex_rows(res.json(), ['UnderlyingStockID', 'SpotID', '標的證券代號'])
            combined_set.update(found)
    except Exception as e:
        print(f"⚠️ [策略一] OpenAPI 獲取異常: {e}")

    # 🛑 [策略二] 官方網頁 Regex 暴力解析 (突破 JS 盾)
    if not combined_set:
        print("⚠️ OpenAPI 遭 WAF 阻擋，啟動 [策略二] 網頁 HTML 暴力解析...")
        try:
            # 直接訪問期交所提供給一般用戶觀看的網頁
            url_html = "https://www.taifex.com.tw/cht/2/stockLists"
            res_html = curl_requests.get(
                url_html, 
                impersonate="chrome120", 
                headers=stealth_headers, 
                verify=False, 
                timeout=10
            )
            
            if res_html.status_code == 200:
                # 使用 Regex 萃取所有 <td> 標籤內的 4 碼數字
                # 匹配格式如：<td align="center">2330</td> 或 <td> 2317 </td>
                import re
                matches = re.findall(r'<td[^>]*>\s*(\d{4})\s*</td>', res_html.text)
                if matches:
                    combined_set.update(matches)
                    print(f"✅ [策略二] 網頁爬蟲成功解析出 {len(matches)} 筆候選代號。")
        except Exception as e:
            print(f"❌ [策略二] 網頁爬蟲備援失敗: {e}")

    # 🛡️ 雜訊過濾：交集台股母體，確保不會混入年份或其他無關數字
    if equity_universe and combined_set:
        combined_set = {sid for sid in combined_set if sid in equity_universe}

    # 💾 寫入快取並回傳
    if combined_set:
        _save_futures_to_cache(combined_set)
        print(f"✅ 最終成功確立 {len(combined_set)} 檔股期名單。")
        return combined_set

    # 🛑 [策略三] 本地快取降級
    cached = _load_futures_from_cache(max_age_days=3)
    if cached:
        print(f"⚠️ [策略三] 線上抓取全數失敗，降級使用本地快取 {len(cached)} 檔（3 天內）。")
        return cached

    print("⚠️ 股期清單獲取失敗，回傳空集合。")
    return set()

# 定義快取檔案路徑
CB_CACHE_FILE = "cb_list_cache.json"

def _is_active_cb(row: dict) -> bool:
    """僅保留目前仍存續的 CB（未到期/未終止/未下市）。"""
    today = get_now_tw().date()

    # 文字狀態先過濾（例如：終止、下市、到期）
    status_text = " ".join(str(v) for k, v in row.items() if any(x in str(k).lower() for x in ['狀態', 'status', '備註', 'remark']))
    if any(x in status_text for x in ['終止', '下市', '到期', '註銷', '償還完畢', 'delist', 'terminated', 'redeemed', 'matured']):
        return False

    # 發行/掛牌日（若尚未生效，視為非活躍）
    issue_date = None
    for k, v in row.items():
        key = str(k).lower()
        if any(x in key for x in ['發行', '掛牌', '上市日', 'issue', 'list']):
            issue_date = _extract_yyyy_mm_dd_from_text(v)
            if issue_date:
                break
    if issue_date and issue_date > today:
        return False

    # 到期/終止/償還日（早於今日則剔除）
    for k, v in row.items():
        key = str(k).lower()
        if any(x in key for x in ['到期', '終止', '償還', '贖回', 'maturity', 'redeem', 'end']):
            exp_date = _extract_yyyy_mm_dd_from_text(v)
            if exp_date:
                return exp_date >= today

    # 若資料沒有可用日期欄，保守視為活躍（避免誤殺）
    return True

def _extract_active_cb_stock_code(row: dict) -> str:
    """萃取 CB 對應現股代號，僅接受可信欄位。"""
    if not _is_active_cb(row):
        return ""

    direct_sid = _extract_sid_by_priority(
        row,
        ['發行公司代號', '公司代號', 'SecuritiesCompanyCode', 'UnderlyingStockID', '標的證券代號']
    )
    if direct_sid:
        return direct_sid

    # 退而求其次：由 CB 債券代號推回前 4 碼現股代號（例如 23301）
    for k, v in row.items():
        key = str(k).lower()
        if any(x in key for x in ['債券代號', '代號', '代碼', 'code', 'bond']):
            s_val = str(v).strip()
            if len(s_val) >= 5 and s_val[:4].isdigit() and s_val[:4] != '0000':
                return s_val[:4]

    return ""

def _load_cb_from_cache(max_age_days: int = 1) -> Set[str]:
    """從本地讀取快取資料，若過期或不存在則回傳空集合"""
    if not os.path.exists(CB_CACHE_FILE):
        return set()
   
    try:
        with open(CB_CACHE_FILE, 'r', encoding='utf-8') as f:
            data = json.load(f)
           
        cache_time = datetime.fromisoformat(data.get("timestamp", "2000-01-01"))
        today = datetime.now(timezone.utc) + timedelta(hours=8)
       
        # 檢查快取是否在有效期限內
        if (today - cache_time).days <= max_age_days:
            return set(data.get("sids", []))
    except Exception as e:
        print(f"⚠️ [快取讀取失敗] 忽略舊快取: {e}")
       
    return set()

def _save_cb_to_cache(cb_set: Set[str]):
    """將成功的抓取結果寫入本地快取"""
    try:
        today = datetime.now(timezone.utc) + timedelta(hours=8)
        data = {
            "timestamp": today.isoformat(),
            "sids": list(cb_set)
        }
        with open(CB_CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception as e:
        print(f"⚠️ [快取寫入失敗] {e}")

def fetch_convertible_bond_set() -> Set[str]:
    """取得目前仍存續（活著）的可轉債對應現股代號集合。"""
    print("📡 正在向證交所與櫃買中心調閱活躍 CB 清單（存續過濾模式）...", flush=True)
    cb_set = set()

    endpoints = [
        "https://openapi.twse.com.tw/v1/opendata/t187ap34_L",  # 證交所
        "https://www.tpex.org.tw/openapi/v1/tpex_bond_main"    # 櫃買中心
    ]

    for url in endpoints:
        try:
            res = curl_requests.get(url, impersonate="chrome120", timeout=15)

            if res.status_code == 200:
                if res.text.strip().startswith("<!DOCTYPE") or "<html" in res.text[:20].lower():
                    print(f"⚠️ [API 警告] 伺服器回傳 HTML，可能遭到防火牆阻擋。URL: {url}")
                    continue

                data = res.json()
                if isinstance(data, list):
                    for row in data:
                        sid = _extract_active_cb_stock_code(row)
                        if sid:
                            cb_set.add(sid)

        except Exception as e:
            print(f"❌ [API 連線異常] URL: {url}, Error: {e}")

        time.sleep(1)

    equity_universe = set(global_volume_lookup.keys()) | set(finmind_industry_map.keys())
    if equity_universe:
        cb_set = {sid for sid in cb_set if sid in equity_universe}

    if cb_set:
        _save_cb_to_cache(cb_set)
        return cb_set

    cached = _load_cb_from_cache(max_age_days=3)
    if cached:
        print(f"⚠️ [CB 清單] 本輪抓取失敗，使用快取 {len(cached)} 檔（3 天內）。")
        return cached

    print("⚠️ [CB 清單] 本輪與快取皆無可用資料，回傳空集合。")
    return set()

def load_official_warrant_targets() -> List[str]:
    print("📡 正在解析全市場權證發行清單 (TWSE官方映射 + 雙因子排序)...", flush=True)
    valid_underlying_sids = set()
   
    name_to_sid = {}
    for sid, info in finmind_industry_map.items():
        clean_name = info['name'].replace('*', '').strip().replace('臺', '台')
        name_to_sid[clean_name] = sid
       
    try:
        res = curl_requests.get(
            "https://openapi.twse.com.tw/v1/opendata/t187ap37_L",
            impersonate="chrome120",
            timeout=50
        )
       
        if res.status_code == 200:
            data = res.json()
            for row in data:
                found_sid = False
               
                # 策略 A: 無結構暴力掃描 (Schema-less)
                for key, val in row.items():
                    str_val = str(val).strip()
                    if len(str_val) == 4 and str_val.isdigit():
                        if str_val in finmind_industry_map or str_val in global_volume_lookup:
                            valid_underlying_sids.add(str_val)
                            found_sid = True
                            break
               
                # 策略 B: 如果數字找不到，掃描名稱進行反向映射
                if not found_sid:
                    for key, val in row.items():
                        sname = str(val).replace('*', '').strip().replace('臺', '台')
                        if sname in name_to_sid:
                            valid_underlying_sids.add(name_to_sid[sname])
                            break
        else:
            print(f"⚠️ [TWSE 權證清單] 伺服器異常，狀態碼: {res.status_code}")
           
    except Exception as e:
        print(f"❌ [TWSE 權證清單解析失敗] {e}")

    warrant_candidates = []
    for sid in valid_underlying_sids:
        if sid not in global_volume_lookup or sid not in global_turnover_lookup:
            continue
           
        today_vol = global_volume_lookup[sid]
        today_turnover = global_turnover_lookup[sid]
       
        if today_vol > 0 and today_turnover > 0:
            warrant_candidates.append({
                'sid': sid,
                'vol': today_vol,
                'turnover': today_turnover
            })

    if warrant_candidates:
        warrant_candidates.sort(key=lambda x: x['turnover'], reverse=True)
        for idx, c in enumerate(warrant_candidates):
            c['rank_t'] = idx
           
        warrant_candidates.sort(key=lambda x: x['vol'], reverse=True)
        for idx, c in enumerate(warrant_candidates):
            c['rank_v'] = idx
           
        warrant_candidates.sort(key=lambda x: x['rank_t'] + x['rank_v'])

    clean_hot_sids = [item['sid'] for item in warrant_candidates][:Config.WARRANT_QUOTA]
    print(f"✅ 權證熱門榜精準解析完成！從 {len(valid_underlying_sids)} 檔標的中篩出 {len(clean_hot_sids)} 檔頂級熱門標的。")
   
    return clean_hot_sids

def _fetch_volume_from_finmind_fallback():
    prev_date = get_previous_trading_day()
    print(f"⚠️ TWSE/TPEX 即時資料為空（可能為週一盤前），嘗試取前一交易日 ({prev_date}) 資料...", flush=True)
    try:
        from FinMind.data import DataLoader
        dl = DataLoader()
        if Config.FINMIND_TOKEN: dl.login_by_token(Config.FINMIND_TOKEN)
        df = dl.taiwan_stock_daily(start_date=str(prev_date), end_date=str(prev_date))
        if df.empty: return
           
        for _, row in df.iterrows():
            sid = str(row['stock_id']).strip()
            if len(sid) == 4:
                vol = int(row.get('Trading_Volume', 0)) // 1000
                turnover = float(row.get('Trading_money', 0))
                if vol > 0:
                    global_volume_lookup[sid] = vol
                    global_turnover_lookup[sid] = turnover
    except Exception as e:
        print(f"❌ FinMind 前一交易日資料取得失敗: {e}")

def pre_market_initialization():
    global global_volume_lookup, global_turnover_lookup, _stocks_with_futures, _stocks_with_cb, _cb_primary_market_map, _tide_analyzer
    twse_sids = set()
    disposition_set = fetch_disposition_stocks()

    print("📡 正在拉取全市場即時量能快照...", flush=True)
    try:
        twse_data = curl_requests.get("https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL", impersonate="chrome120", timeout=10).json()
        if isinstance(twse_data, list):
            for row in twse_data:
                code = row.get('Code', '').strip()
                if len(code) == 4:
                    global_volume_lookup[code] = safe_cast(row.get('TradeVolume'), int) // 1000
                    global_turnover_lookup[code] = safe_cast(row.get('TradeValue'), float)
                    twse_sids.add(code)
    except: pass
    try:
        tpex_data = curl_requests.get("https://www.tpex.org.tw/openapi/v1/tpex_mainboard_quotes", impersonate="chrome120", timeout=10).json()
        if isinstance(tpex_data, list):
            for row in tpex_data:
                code = str(row.get('SecuritiesCompanyCode', '')).strip()
                if len(code) == 4:
                    global_volume_lookup[code] = safe_cast(row.get('TradingShares'), int) // 1000
                    global_turnover_lookup[code] = safe_cast(row.get('TradingAmount'), float)
    except: pass

    if len(global_volume_lookup) == 0:
        _fetch_volume_from_finmind_fallback()

    _stocks_with_futures = fetch_stock_futures_set()
    _stocks_with_cb = fetch_convertible_bond_set()
    _cb_primary_market_map = load_cb_primary_market_map_from_xlsx()
    _tide_analyzer = TideAnalyzer(_resolve_tide_concepts_path())
    print(f"📊 衍生商品偵測完成：股期 {len(_stocks_with_futures)} 檔、CB {len(_stocks_with_cb)} 檔。")

    clean_sids = load_official_warrant_targets()

    injected = 0
    for sid in clean_sids:
        if sid in disposition_set: continue
        ind = finmind_industry_map.get(sid, {}).get('industry', '權證熱門標的')
        name = finmind_industry_map.get(sid, {}).get('name', f"現股_{sid}")
        if should_exclude(sid, name, ind): continue
       
        if sid not in stock_info_map:
            y_vol_base = global_volume_lookup.get(sid, 1500)
            if y_vol_base < 100: y_vol_base = 1500
            _exchange_map[sid] = 'tse' if sid in twse_sids else 'otc'
            stock_info_map[sid] = {'name': name, 'market': '權證', 'is_protected': True, 'industry': ind}
            monitor_data[sid] = {
                "high": 0.0, "low": 9999.0, "y_vol": y_vol_base,
                "state": 0, "point_a": -1.0, "point_b": 9999.0,  
                "last_alert_time": 0, "last_up_pct": 0.0, "last_ratio": 1.0, "last_consumption": 0.85, "is_accelerating": True, "history_prices": [],
                "candle_window": [], "last_alert_by_strategy": {}
            }
            injected += 1
            if injected >= Config.WARRANT_QUOTA: break
           
    print(f"🎫 動態權證現股化分析完成，最終注入 {injected} 檔實時金流標的。")

def update_rolling_3k(sid, lp):
    data = monitor_data[sid]
    now_ts = time.time()
    bucket_sec = Config.CANDLE_TIMEFRAME_MIN * 60
    candle_start_ts = int(now_ts // bucket_sec) * bucket_sec

    cw = data.setdefault('candle_window', [])
    if cw and cw[-1][0] == candle_start_ts:
        cw[-1][1] = max(cw[-1][1], lp)
        cw[-1][2] = min(cw[-1][2], lp)
    else:
        cw.append([candle_start_ts, lp, lp])
        if len(cw) > Config.CANDLE_3K_COUNT:
            cw.pop(0)

    if cw:
        data['high'] = max(c[1] for c in cw)
        data['low']  = min(c[2] for c in cw)

def recover_3k_data(target_list: List[str]):
    now_tw = get_now_tw()
    if not is_market_hours(): return
   
    today_tw = now_tw.date()
    bucket_sec = Config.CANDLE_TIMEFRAME_MIN * 60
    current_candle_start_ts = int(now_tw.timestamp() // bucket_sec) * bucket_sec
    print(f"🔄 執行 3K 補課 (共 {len(target_list)} 檔，取最近 {Config.CANDLE_3K_COUNT} 根 {Config.CANDLE_TIMEFRAME_MIN}分K)...", flush=True)
   
    for sid in target_list:
        try:
            url = f"https://api.fugle.tw/marketdata/v1.0/stock/intraday/candles/{sid}?timeframe={Config.CANDLE_TIMEFRAME_MIN}"
            res = standard_requests.get(url, headers={"X-API-KEY": Config.FUGLE_API_KEY.strip()}, timeout=6).json()
            if res and "data" in res:
                kbars = res.get('data', [])
                today_bars = []
                for k in kbars:
                    k_dt = isoparse(k['date']).astimezone(timezone(timedelta(hours=8)))
                    bar_start_ts = int(k_dt.timestamp() // bucket_sec) * bucket_sec
                    if k_dt.date() == today_tw and bar_start_ts < current_candle_start_ts:
                        today_bars.append([bar_start_ts, k['high'], k['low']])
                today_bars.sort(key=lambda x: x[0])
                last3 = today_bars[-Config.CANDLE_3K_COUNT:]
                if last3:
                    monitor_data[sid]['candle_window'] = last3
                    monitor_data[sid]['high'] = max(c[1] for c in last3)
                    monitor_data[sid]['low']  = min(c[2] for c in last3)
        except: pass
        time.sleep(Config.API_THROTTLE_SLEEP)

# ------------------------------------------------------------
# 🎬 主程式核心驅動流
# ------------------------------------------------------------

def main():
    global _last_fugle_scan, finmind_industry_map
    print(f"🛡️ 蘇蘇的天機選股 V2.3 啟動完成。")
    print(f"{get_now_tw().strftime('%H:%M:%S')} 執行盤前籌碼映射與官方 Profile 基本面同步...")
   
    finmind_industry_map = fetch_finmind_industry_mapping()
    pre_market_initialization()
   
    print("🛰️ 啟動市場動態配額同步模組...")
    refresh_pool_v90()
   
    w_count = len([s for s in stock_info_map if stock_info_map[s]['market'] == '權證'])
    l_count = len([s for s in stock_info_map if stock_info_map[s]['market'] == '上市'])
    o_count = len([s for s in stock_info_map if stock_info_map[s]['market'] == '上櫃'])
    print(f"✅ 初始化：上市 {l_count} 檔、上櫃 {o_count} 檔、權證主力(protected) {w_count} 檔。")
   
    perform_strategy_test()
    recover_3k_data(list(stock_info_map.keys()))
    print("🚀 系統初始化完畢，準備進入監控模式...\n")
    time.sleep(0.5)
   
    max_results = {}
   
    while True:
        refresh_pool_v90()
       
        _need_recover = [s for s in stock_info_map if monitor_data.get(s, {}).get('high', 0.0) == 0.0]
        if _need_recover and is_market_hours():
            recover_3k_data(_need_recover[:15])
           
        tw_now = get_now_tw()
       
        if not Config.IS_LOCAL:        
            if tw_now.weekday() < 5 and tw_now.time() >= Config.AUTO_SHUTDOWN_TIME:
                print(f"\n[系統時鐘觸發熄火] 當前台北時間 {tw_now.strftime('%H:%M:%S')} 已達 13:35。")
                sys.exit(0)

        timer_str = tw_now.strftime('%Y-%m-%d %H:%M:%S')
        print(f"\n[監控週期: {timer_str}]")
        print(f"{wide_ljust('股號股名', 20)} | {wide_ljust('市場', 6)} | {wide_ljust('現價', 10)} | {wide_ljust('量比', 8)} | {wide_ljust('3K高', 10)} | {wide_ljust('3K低', 10)} | 衍生品")
        print("-" * 115)
       
        sorted_sids = sorted(stock_info_map.keys(), key=lambda x: 0 if stock_info_map[x]['is_protected'] else (1 if stock_info_map[x]['market'] == '上市' else 2))
       
        passed_min = (datetime.combine(tw_now.date(), tw_now.time()) - datetime.combine(tw_now.date(), Config.MARKET_OPEN)).total_seconds() / 60
        if passed_min <= 0 or passed_min > 270: passed_min = 270.0
           
        mis_ok = False
        if any(not stock_info_map[s]['is_protected'] for s in sorted_sids):
            max_results = fetch_mis_batch_all()
            mis_ok = len(max_results) > 0
            if mis_ok: print(f"[快速層] MIS 實時雷達運作正常，已捕獲 {len(max_results)} 檔即時行情快照。")

        for sid in sorted_sids:
            info, data = stock_info_map[sid], monitor_data[sid]
            try:
                lp, v, up_pct = None, 0, 0.0
                lp_is_fresh = False

                if info.get('is_protected'):
                    f_url = f"https://api.fugle.tw/marketdata/v1.0/stock/intraday/quote/{sid}"
                    res = standard_requests.get(f_url, headers={"X-API-KEY": Config.FUGLE_API_KEY.strip()}, timeout=5).json()
                    if res and 'lastPrice' in res and res['lastPrice'] is not None:
                        lp = res.get('lastPrice')
                        lp_is_fresh = True
                        v = safe_cast(res.get('total', {}).get('tradeVolume'), int)  
                        ask_vol = sum(safe_cast(a.get('volume', 0), int) for a in res.get('asks', [])[:3])
                        if ask_vol > 0: data['last_consumption'] = min(1.0, v / (ask_vol * 10))
                       
                        y = safe_cast(res.get('previousClose', '0'), float)
                        if y > 0: up_pct = round((lp - y) / y * 100, 2)
                        data['last_up_pct'] = up_pct
                else:
                    if mis_ok and sid in max_results:
                        m = max_results[sid]
                        lp, v = m['lp'], m['v']
                        lp_is_fresh = m.get('is_traded', False)
                        up_pct = m.get('up_pct', 0.0)
                        if m.get('ask3', 0) > 0: data['last_consumption'] = min(1.0, v / (m['ask3'] * 10))
                        data['last_up_pct'] = up_pct

                if lp is None and Config.IS_LOCAL:
                    y_val = safe_cast(finmind_industry_map.get(sid, {}).get('y', '0'), float)
                    seed_base = y_val if y_val > 0 else (1000.0 if sid=='2330' else (210.0 if sid=='2317' else 85.0))
                    wave = math.sin(passed_min * 0.1 + int(sid)) * 0.05
                    lp = round(seed_base * (1.002 + wave), 2)
                    v = int(data['y_vol'] * (passed_min / 270.0) * (1.1 + abs(wave)))
                    lp_is_fresh = True

                if lp is None:
                    if data.get('history_prices'): lp = data['history_prices'][-1]
               
                if not lp: continue
               
                _old_3k_high = data['high']
                update_rolling_3k(sid, lp)

                if 'history_prices' not in data: data['history_prices'] = []
                data['history_prices'].append(lp)
                if len(data['history_prices']) > 5: data['history_prices'].pop(0)
                data['is_accelerating'] = data['history_prices'][-1] > data['history_prices'][-2] if len(data['history_prices']) >= 2 else False
               
                ratio = round((v * (270 / passed_min)) / data['y_vol'], 2) if data['y_vol'] > 0 else 0
                data['last_ratio'] = ratio

                stock_label = f"{sid} {info.get('name', '')}"
                market_label = info['market']
                fut_flag = '✅' if sid in _stocks_with_futures else '❌'
                cb_flag = '✅' if sid in _stocks_with_cb else '❌'
                deriv_flag = f"期{fut_flag}|債{cb_flag}"
               
                print(f"{wide_ljust(stock_label, 20)} | {wide_ljust(market_label, 6)} | {wide_ljust(lp, 10)} | {wide_ljust(ratio, 8)} | {wide_ljust(data['high'], 10)} | {wide_ljust(data['low'], 10)} | {deriv_flag}")

                if not lp_is_fresh:
                    if Config.IS_LOCAL and info.get('is_protected'):
                        time.sleep(Config.API_THROTTLE_SLEEP)
                    continue

                is_3k_break = lp > _old_3k_high > 0
                is_vol_anomaly = ratio >= Config.VOL_EST_THRESHOLD

                if data['point_a'] < 0:
                    data['point_a'] = lp
                else:
                    _n_valid_pullback = data['point_b'] != 9999.0 and data['point_b'] <= data['point_a'] * 0.985
                   
                    if data['state'] == 1 and lp >= data['point_a'] and _n_valid_pullback:
                        if is_vol_anomaly:
                            send_tg_alert(sid, "策略四：N字突破 (洗盤結束再發動)", lp, data['high'], data['low'], ratio, up_pct)
                            data['point_a'], data['point_b'], data['state'] = lp, 9999.0, 0
                    elif lp > data['point_a']:
                        data['point_a'], data['point_b'], data['state'] = lp, 9999.0, 1
                    elif data['state'] == 1 and lp < data['point_a']:
                        data['point_b'] = min(data['point_b'], lp)
                        if lp < data['point_a'] * 0.93:
                            data['point_b'] = 9999.0
                            data['state'] = 0

                if is_3k_break and is_vol_anomaly:
                    send_tg_alert(sid, "🔥 策略二：3K突破 + 量能異常 (價量齊揚)", lp, _old_3k_high, data['low'], ratio, up_pct)
                    # ==========================================
                    # 🚀 新增：LINE 權證主力專屬推播邏輯
                    # ==========================================
                    if info.get('is_protected'):                        
                        tw_now = get_now_tw()
                        trigger_time = tw_now.strftime('%H:%M:%S')
                            
                        # 參數映射與格式轉換
                        # 壓力消化：取用你前面算好的 last_consumption (五檔內盤消耗比)
                        pressure = f"{data.get('last_consumption', 0.0):.1%}"
                        # 能量斜率：取用你算好的 is_accelerating (連續兩次價格上升)
                        slope = "📈 動能上升" if data.get('is_accelerating') else "平緩/震盪"
                        # 產業別
                        ind = info.get('industry', '未分類')
                        
                        # 🛡️ 冷卻檢查：30 分鐘內同策略不重複通知 (與 Telegram 同步)
                        strategy_name = "🔥 策略：3K突破量能異常"
                        alert_times = data.setdefault('last_alert_by_strategy', {})
                        if time.time() - alert_times.get(strategy_name, 0) >= 1800:
                            send_line_flex_warrant_alert(
                                sid=sid,
                                name=info.get("name", ""),
                                strategy=strategy_name,
                                lp=lp,
                                pct_str=f"{up_pct:+.2f}%",
                                up_pct=up_pct,
                                ratio=ratio,
                                stop_loss=_old_3k_high,  # 實戰中常以 3K 高點作為突破後的防守價
                                pressure_digestion=pressure,
                                energy_slope=slope,
                                fut_flag=fut_flag,
                                cb_flag=cb_flag,
                                industry=ind,
                                time_str=trigger_time
                            )
                            alert_times[strategy_name] = time.time()
                            print(f"✅ [LINE] 成功派發權證主力通知: {sid} {info.get('name', '')}")                        
                    # ==========================================  
                if info.get('is_protected'): time.sleep(Config.API_THROTTLE_SLEEP)
            except: pass
          


        if time.time() - _last_fugle_scan >= 60 and len(stock_info_map) > 0:
            protected_sids = [s for s in stock_info_map if stock_info_map[s].get('is_protected')]
            if protected_sids:
                print(f"\n[慢速層] 啟動 Fugle 實時精確五檔掛單分析 ({len(protected_sids)} 檔)...")
                for sid in protected_sids:
                    try:
                        f_url = f"https://api.fugle.tw/marketdata/v1.0/stock/intraday/quote/{sid}"
                        res = standard_requests.get(f_url, headers={"X-API-KEY": Config.FUGLE_API_KEY.strip()}, timeout=5).json()
                        if res and 'lastPrice' in res and res['lastPrice'] is not None:
                            ask_vol = sum(safe_cast(a.get('volume', 0), int) for a in res.get('asks', [])[:3])
                            v_fugle = safe_cast(res.get('total', {}).get('tradeVolume'), int)
                            if ask_vol > 0:
                                monitor_data[sid]['last_consumption'] = min(1.0, v_fugle / (ask_vol * 10))
                           
                            w_lp = monitor_data[sid].get('history_prices', [100.0])[-1]
                            warrant_label = f"{sid} {stock_info_map[sid].get('name', '')}"
                            print(f"{wide_ljust(warrant_label, 20)} | {wide_ljust('權證', 6)} | {wide_ljust(w_lp, 10)} | {wide_ljust(monitor_data[sid].get('last_ratio', 1.0), 8)} | {wide_ljust(monitor_data[sid]['high'], 10)} | {wide_ljust(monitor_data[sid]['low'], 10)} | {stock_info_map[sid]['industry']}")
                    except: pass
                    time.sleep(Config.API_THROTTLE_SLEEP)
            _last_fugle_scan = time.time()
            print("")
           
        print(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
        print(f"[時段提示] 方案C 兩層架構穩定運行 (MIS {len(max_results)} 檔)。5秒後刷新...")
        time.sleep(5)

if __name__ == "__main__": main()
